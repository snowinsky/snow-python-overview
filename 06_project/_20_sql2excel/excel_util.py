import openpyxl as x_open



class excel_reader(object):
    def __init__(self, f=None):
        self.xlsx_file = f

    def read_sheet(self):
        pass

    def read_line(self):
        pass

class excel_writer(object):
    def __init__(self, f=None):
        self.xlsx_file = f

    def write(self):
        pass


wb = x_open.load_workbook(xlsx_path)
ws1 = wb['Sheet1']
print(ws1['A1'])
ws1['A1'] = 42
ws1['A2'] = '042'

for row in ws1.rows:
    cellValue = [];
    for cell in row:
        cellValue.append(str(cell.value))
    print(',   '.join(cellValue))

sheetNameList = wb.sheetnames
print(sheetNameList)
if ('Sheet2' in sheetNameList):
    ws2 = wb['Sheet2']
else:
    ws2 = wb.create_sheet('Sheet2')

for x in range(10):
    ws2['A' + str(x + 1)] = x

print(wb.chartsheets)

wb.save(xlsx_path)